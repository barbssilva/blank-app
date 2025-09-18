import openpyxl
import time
import sys
import pandas as pd
import warnings
from openpyxl.styles import Alignment, Border, Side
import numpy as np
from openpyxl import load_workbook
import os
from rapidfuzz import process, fuzz
import re
import copy

warnings.simplefilter("ignore", UserWarning)

# Função para converter strings do estilo "26,9" em números, mantendo NaN
def converter_para_float(valor):
    try:
        if pd.isna(valor):  # Manter NaN
            return np.nan
        if isinstance(valor, (int, float)):  # Se já for número, retornar como float
            return float(valor)
        valor = str(valor).strip().replace(',', '.')  # Substituir vírgula por ponto
        return float(valor)
    except:
        return np.nan  # Se não conseguir converter, retorna NaN

"""
SELECIONAR SHEETS PARA TRADUZIR
"""

def escolher_sheets(excel_path,excel_output1,keywords1):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_path,data_only=True)
    sheet_names = workbook.sheetnames

    # Filtrar sheets a manter
    sheets_to_keep = [sheet_name for sheet_name in sheet_names if any(keyword.lower() in sheet_name.strip().lower() for keyword in keywords1)]
    sheets_to_remove = [sheet_name for sheet_name in sheet_names if sheet_name not in sheets_to_keep]

    # Remover sheets não desejadas
    for sheet_name in sheets_to_remove:
        workbook.remove(workbook[sheet_name])
        
    # Guardar o ficheiro com tradução
    workbook.save(excel_output1)
    return


#esta função serve para desunir as celulas da primeira coluna, para termos espaço para inserir a tradução sem estragar a
# formatação da restante tabela
def preparar_celulas_traducao(excel_path, linha_inicio=6):
    wb = openpyxl.load_workbook(excel_path)

    #desunir primeira coluna
    for sheet in wb.worksheets:
        sheet_name_lower = sheet.title.lower()
        if not ("spec" in sheet_name_lower or "grading" in sheet_name_lower):
            continue  # Ignora sheets que não têm 'spec' ou 'grading' no nome

        # Encontrar todas as células unidas que começam na primeira coluna e linha >= linha_inicio
        for merged_range in list(sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_col == 1 and min_row >= linha_inicio:
                sheet.unmerge_cells(str(merged_range))

        # Definir largura das colunas antes de unir
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        #unir A-B e C-D
        max_row = sheet.max_row
        for row in range(linha_inicio, max_row + 1):
            # Unir A e B
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            # Unir C e D
            sheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)        

    wb.save(excel_path)
    wb.close()
    return

def traducao(excel_output1):
    traducao_dict={"armhole straight":"CAVA A DIREITO",
        "back neck drop":"PROF. DECOTE COSTAS",
        "back neck width - straight":"ABERTURA DECOTE COSTAS",
        "back neck width - snp to snp straight - straight":"ABERTURA DECOTE COSTAS",
        "back rise - including waistband":"GANCHO COSTA INCL. CINTO",
        "back pocket width":"LARGURA BOLSO COSTA",
        "back pocket depth":"ALTURA BOLSO COSTA",
        "back pocket position down from waistband to cb edge":"POSIÇÃO BOLSO DESDE CENTRO COSTA",
        "back pocket position down from waistband to ss edge":"POSIÇÃO BOLOSO DESDE CINTO",
        "back armhole curved":"CAVA COSTA CURVADA",
        "belt length":"COMPRMENTO DO CINTO",
        "belt depth":"ALTURA DO CINTO",
        "cb length - cb back neck seam to hem edge":"COMP. COSTA DESDE DECOTE COSTA",
        "cuff depth":"ALTURA PUNHO",
        "collar depth at cb":"ALTURA GOLA",
        "exposed drawcord length":"CORDÃO EXPOSTO",
        "front length - snp to bottom hem edge":"COMP. FRENTE DESDE PO",
        "front neck drop - snp to cf invisible line":"PROF.DECOTE FRENTE DESDE PO",
        "front channel hood":"LARGURA ENTRE CASAS",
        "front rise - including waistband":"GANCHO FRENTE INCL.CINTO",
        "front pocket width -along waistband seam":"POSIÇÃO BOLSO NO CINTO",
        "front pocket depth - measure along side seam":"ABERTURA BOLSO",
        "front neck curved":"DECOTE FRENTE CURVADO",
        "front armhole curved":"CAVA FRENTE CURVADA",
        "hem width":"FUNDO",
        "hem width measured at peak of hem curve":"FUNDO",
        "hem width on seam":"FUNDO NA COSTURA",
        "hem width on rib edge":"FUNDO NA COSTURA",
        "hem depth":"ALTURA BAINHA FUNDO",
        "hem width on bottom edge":"ABERTURA PERNA",
        "hood height - cf neck to top of hood":"ALTURA CAPUZ - DESDE CENTRO COSTA",
        "hood dart length":"COLOCAÇÃO DAS CASAS AO DECOTE",
        "hood height along cf edge":"ALTURA DO CAPUZ NO CENTRO FRENTE",
        "hood length along cb edge":"ALTURA DO CAPUZ NO CENTRO COSTA",
        "hood width at widest":"LARGURA DO CAPUZ NA PARTE MAIS LARGA",
        "inside leg length":"PERNA POR DENTRO",
        "lower pocket width at top edge - total":"LARGURA DO BOLSO NO TOPO",
        "lower pocket width at hem seam":"LARGURA BOLSO NO FUNDO",
        "lower pocket width at widest":"LARGURA BOLSO NA PARTE MAIS LARGA",
        "lower pocket at centre":"ALTURA BOLSO NO CENTRO",
        "lower pocket depth at side seam":"ALTURA BOLSO NA COSTURA LATERAL",
        "lower pocket width - along top edge":"LARGURA DO BOLSO FRENTE NO TOPO",
        "lower pocket width - bottom edge":"LARGURA DO BOLSO FRENTE NO FUNDO",
        "lower pocket length - at pocket side":"ALTURA DO BOLSO NAS LATERAIS",
        "lower pocket depth - through centre":"ALTURA DO BOLSO NO CENTRO",
        "min neck stretch - minimum":"DECOTE MINIMO ESTICADO",
        "nape to cuff":"NUCA AO PUNHO",
        "outer leg -waist seam to hem":"PERNA POR FORA EXCL.CINTO",
        "overhood length cb to cf inc trim": "COMPR. CAPUZ DSD CENTRO COSTA AO CENTRO FRENTE",
        "print position from cf neckline":"POSIÇÃO ESTAMPADO DESDE DECOTE FRENTE",
        "pocket opening":"ABERTURA DO BOLSO",
        "shoulder drop - from invisible line to shoulder point":"QUEDA DE OMBRO",
        "shoulder at seam / fold":"OMBRO",
        "shoulder to shoulder":"OMBRO A OMBRO",
        "sleeve length - along outside edge":"COMPRIMENTO DA MANGA",
        "side slit length":"COMPRIMENTO DA FENDA NA LATERAL",
        "visible drawcord length": "CORDÃO VÍSIVEL",
        "waist width from hsp":"CINTA DESDE PO",
        "waist width":"CINTA",
        "waist width at top edge - edges together":"CINTA NO TOPO",
        "waist band depth":"ALTURA CINTO",
        "waist position - below shoulder neck point":"POSIÇÃO DA CINTA - ABAIXO DO PO",
        "zip length":"COMPRIMENTO DO FECHO",
        "bust - measured at {num}cm below underarm": "PEITO A {num}CM DA CAVA",
        "chest-measured at {num}cm below underarm": "PEITO A {num}CM DA CAVA",
        "forearm width {num} below underam":"ANTEBRAÇO A {num}CM DA CAVA",
        "high hip width - {num}cm below waist edge":"ANCA ALTA A {num} CM DA COSTURA DO CINTO",
        "hood width {num}cm down inc trim":"LARGURA CAPUZ A {CM} DAS CASAS",
        "low hip width - {num} cm below high hip":"ANCA BAIXA A {num} CM DA ANCA ALTA",
        "strap length (including {num}cm adjuster)":"COMPR. ALÇA (INCLUÍNDO AJUSTADOR DE {num}CM)",
        "thigh width - at {num}cm below crotch":"COXA A {num} CM DO GANCHO",
        "x front - {num}cm below shoulder neck point":"ENTRE CAVAS FRENTE A {num}CM DO PO",
        "x back - {num}cm below shoulder neck point":"ENTRE CAVAS COSTA A {num}CM DO PO",
        "1/2 bicep width -measured at {num}cm below underarm":"MÚSCULO A {num}CM DA CAVA",
        "1/2 bicep width at {num}cm down from underarm":"MÚSCULA A {num}CM DA CAVA",
        "1/2 elbow width - {num}cm down from underarm":"COTOVELO A {num}CM DA CAVA",
        "1/2 high hip - {num}cm below waist":"ANCA ALTA A {num} CM ABAIXO DA CINTA",
        "1/2 low hip - {num}cm below waist":"ANCA BAIXA A {num} CM ABAIXO DA CINTA",
        "raise {num] cm": "SUBIR {num}CM",
        "1/2 cuff width":" LARGURA DO PUNHO",
}
    
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(excel_output1)
    sheet_names = workbook.sheetnames
    referencias_guardadas = {}

    # 1️⃣ Guardar fórmulas para restaurar depois
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):  
                    referencias_guardadas[(sheet.title, cell.coordinate)] = cell.value  

    # 2️⃣ Guardar textos a partir da linha 5 em cada sheet
    textos_para_traduzir = set()

    sheets_para_traduzir = [sheet_name for sheet_name in sheet_names if any(keyword.lower() in sheet_name.strip().lower() for keyword in ['spec', "grading"])]
    for sheet in workbook.worksheets:
        if sheet.title not in sheets_para_traduzir:
            continue


        print(f"A traduzir: {sheet.title}")
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
            if row_idx < 6:
                continue
            cell = row[0]  # Ver elemento que está apenas na primeira coluna 
            for cell in row:
                if isinstance(cell.value, str) and not cell.value.startswith("=") and cell.value.strip():
                    # Verificar se o conteúdo não é composto apenas por símbolos (como - ou ?)
                    if not re.match(r'^[^\w\s]+$', cell.value.strip()):
                        textos_para_traduzir.add(cell.value.strip())
                       

        comparar_traducoes = list(traducao_dict.keys())
        traducoes = {}

        for frase in textos_para_traduzir:
            frase_lower = frase.lower()
            melhor_match, score, indice = process.extractOne(frase_lower, comparar_traducoes, scorer=fuzz.WRatio)
            match = re.search(r"(\d+(?:\.\d+)?)\s*[cC][mM]", frase)
            if match:
                if score >= 85:
                    template_traducao = list(traducao_dict.values())[indice]
                    valor = match.group(1)
                    if '{num}' in template_traducao:
                        resultado = template_traducao.replace('{num}', valor)
                        traducoes[frase] = resultado
                    else:
                        traducoes[frase] = ""
                else:
                    traducoes[frase] = ""
            else:
                if score >= 89:
                    template_traducao = list(traducao_dict.values())[indice]
                    traducoes[frase] = template_traducao
                else:
                    traducoes[frase] = ""


        # 3️⃣ Preencher a tradução na coluna 3 (C) de cada linha
        for sheet in workbook.worksheets:
            if sheet.title not in sheets_para_traduzir:
                continue
            for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                if row_idx < 6:
                    continue
                cell = row[0]
                #verificar se o valor da célula não é fórmula, não está vazio e não é uma formula (isto é não começa por igual)
                if isinstance(cell.value, str) and not cell.value.startswith("=") and cell.value.strip():
                    #garantir que o conteudo da celula não é apenas simbolos, tipo -, ?, etc
                    if not re.match(r'^[^\w\s]+$', cell.value.strip()):
                        #procurar a tradução no dicionário, removendo espaços extra no texto, e devolve string "" se não encontrar tradução
                        traducao = traducoes.get(cell.value.strip(), "")
                        #introduzir tradução na mesma linha mas na coluna 3 (C)
                        sheet.cell(row=row_idx, column=3).value = traducao

    workbook.save(excel_output1)
    return traducoes


def add_tabelas_traducoes(excel_path, excel_output2, keywords,traducoes):
    # Ler todas as sheets do arquivo Excel
    df_dict = pd.read_excel(excel_path, sheet_name=None,dtype=str)
    #Filtrar as sheets com base nos nomes que contêm determinadas palavras (case-insensitive)
    filtered_sheets = {sheet_name: df for sheet_name, df in df_dict.items()
                    if any(keyword.lower() in sheet_name.strip().lower() for keyword in keywords)}

    # Salvar as sheets filtradas em um novo arquivo Excel
    with pd.ExcelWriter(excel_output2, engine='xlsxwriter') as writer:
        for sheet_name, df in filtered_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


    #DEPOIS DE CRIAR O FICHEIRO CALCULAR A DIFERENÇA ENTRE OS TAMANHOS
    excel_data = pd.read_excel(excel_output2, sheet_name=None)
    # Processar cada sheet
    for sheet_name, df in excel_data.items():
        # Encontrar a linha que contém "Tol +/-"
        size_row = None
        for index, row in df.iterrows():
            cleaned_values = row.astype(str).str.strip()  # Remover espaços extras das células
            if any(cleaned_values == "Tol +/-"):  # Verificar se "Tol +/-" está presente
                size_row = index
                break

        # Se encontrar a linha, eliminar todas as linhas antes dela e usar como cabeçalho
        if size_row is not None:
            df = df.iloc[size_row:]  # Manter as linhas a partir de size_row
            df.columns = df.iloc[0].astype(str).str.strip()  # Limpar espaços extras no cabeçalho
            df = df.drop(df.index[0]).reset_index(drop=True)  # Remove a linha que agora é o cabeçalho

        # Atualizar o DataFrame no dicionário
        excel_data[sheet_name] = df
    
    # Salvar as alterações no próprio arquivo Excel (sobrescrever o arquivo original)
    with pd.ExcelWriter(excel_output2, mode='w') as writer:  # mode='w' sobrescreve o arquivo original
        for sheet_name, df in excel_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Passo 3: Ler novamente o arquivo atualizado
    excel_data = pd.read_excel(excel_output2, sheet_name=None)
    #calcular a diferença entre os tamanhos
    SIZES = {'UK2', 'UK4', 'UK6', 'UK8', 'UK10', 'UK12', 'UK14', 'UK16', "XXS", "XS", "S", "M", "L", "XL", "XXL", "XXXL"
             'XS/28','S/30','M/32','L/34','XL/36','XXL/38', 'XXS/34','XS/34','S/36','M/38','L/40','XL/42','XXL/44','XXXL'}


    colunas_diferenca=[]
    for sheet_name, df in excel_data.items():
        # Selecionar apenas as colunas que contêm tamanhos específicos
        size_columns = [col for col in df.columns.tolist() if any(str(col).replace(" ", "") == str(size).replace(" ", "") for size in SIZES)]
        for i in range(len(size_columns)):
            df.loc[:, size_columns[i]] = df[size_columns[i]].apply(converter_para_float)

        for i in range(1, len(size_columns)):
            coluna_atual = size_columns[i]
            coluna_anterior = size_columns[i - 1]
                
            # Nome da nova coluna de diferença
            nova_coluna = f'Dif {coluna_atual}-{coluna_anterior}'
            colunas_diferenca.append(nova_coluna)
                
            # Calcular a diferença
            df.loc[:, nova_coluna] = df.loc[:, coluna_atual] - df.loc[:, coluna_anterior]
                
            # Encontrar o índice da coluna atual para inserir logo após ela
            indice_coluna = df.columns.get_loc(coluna_atual)
            colunas_atualizadas = df.columns.tolist()
                
            # Remover a nova coluna do final e inserir após a coluna atual
            colunas_atualizadas.remove(nova_coluna)
            colunas_atualizadas.insert(indice_coluna, nova_coluna)
                
            # Atualizar a ordem das colunas no DataFrame
            df = df.loc[:, colunas_atualizadas]
                
            #remover 0s das colunas
            df.loc[:,coluna_atual] = df.loc[:,coluna_atual].replace(0, np.nan)
            df.loc[:,coluna_anterior] = df.loc[:,coluna_anterior].replace(0, np.nan)
            df.loc[:,nova_coluna] = df.loc[:,nova_coluna].replace(0, np.nan)
            

            # Encontra o índice da última coluna da lista de colunas com tamanhos
            last_col_index = df.columns.get_loc(size_columns[-1])
                
            # Mantém todas as colunas até o índice da última coluna da lista com tamanhos
            df= df.iloc[:, :last_col_index + 1]

            #remover nome de colunas unnamed
            #df.columns = ["" if "Unnamed" in str(col) else col for col in df.columns]

            #remover 0s da primeira coluna, que foram adicionados ao ler o ficheiro
            df.iloc[:, 0] = df.iloc[:, 0].replace("0", "") 

            #remover colunas vazias
            df = df.dropna(axis=1, how='all')  # Agora remove as colunas completamente vazias

            #remover do nome das colunas nan e substituir por string vazia
            df.columns = ["" if "nan" in str(col).lower() else col for col in df.columns]

            # Atualizar o DataFrame no dicionário
            excel_data[sheet_name] = df

        # renomear as colunas com diferença entre tamanhos por "":
        novo_nome = {col: "" for col in colunas_diferenca if col in df.columns}
        df = df.rename(columns=novo_nome)
        excel_data[sheet_name] = df
    

        #Salvar as alterações no próprio arquivo Excel (sobrescrever o arquivo original)
        with pd.ExcelWriter(excel_output2, mode='w') as writer:  # mode='w' sobrescreve o arquivo original
            for sheet_name, df in excel_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = openpyxl.load_workbook(excel_output2)
    for sheet in wb.worksheets:
        # Inserir nova coluna B (posição 2)
        sheet.insert_cols(2)
        max_row = sheet.max_row
        for row in range(2, max_row + 1):
            valor_original = sheet.cell(row=row, column=1).value
            if isinstance(valor_original, str) and not valor_original.startswith("=") and valor_original.strip():
                if not re.match(r'^[^\w\s]+$', valor_original.strip()):
                    traducao = traducoes.get(valor_original.strip(), "")
                    sheet.cell(row=row, column=2).value = traducao
    wb.save(excel_output2)
    wb.close()

    return


def formatar_excel(nome_excel):
    workbook = openpyxl.load_workbook(nome_excel)
    for sheet in workbook.worksheets:
        #adicionar "_copy" ao nome de cada sheet para depois poder juntar ao excel original
        if not sheet.title.endswith("_copy"):
            sheet.title = f"{sheet.title}_copy"


        # Definir a largura da primeira coluna (A)
        sheet.column_dimensions['A'].width = 45
        sheet.column_dimensions['B'].width = 45

        start_col_idx = openpyxl.utils.column_index_from_string('C')

        # Iterar por todas as colunas a partir de 'D' e ajustar a largura
        for col_idx in range(start_col_idx, sheet.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 6

        # Ajustar a altura das linhas e fazer o texto se ajustar ao tamanho da célula
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                sheet.row_dimensions[cell.row].height = 15  # Ajuste a altura conforme necessário

        for row in sheet.iter_rows():
            for cell in row:
                if cell.column >= start_col_idx:  # Apenas células a partir da coluna 'E'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula
        
    # Salvar o arquivo Excel
    workbook.save(nome_excel)
    workbook.close()

    wb = load_workbook(nome_excel)
    ws = wb.active
    
    for sheet in wb.worksheets:
        # Iterar sobre as células da primeira linha (cabeçalho) e aplicar wrap_text e centralizar
        for cell in sheet[1]:  # sheet[1] acessa a primeira linha
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Ajustar a altura da primeira linha para garantir que o texto quebre
        sheet.row_dimensions[1].height = 25  # Ajuste conforme necessário

    # Salvar as alterações
    wb.save(nome_excel)
    wb.close()


def add_info(excel_output1, excel_output2):
    
    wb_modelo = openpyxl.load_workbook(excel_output1)
    ws_modelo = wb_modelo.worksheets[0]  # Primeira sheet do modelo

    # Guardar as 3 primeiras linhas (incluindo valores, fórmulas e estilos)
    linhas_modelo = []
    for row in ws_modelo.iter_rows(min_row=1, max_row=3, max_col=2):
        linha = []
        for cell in row:
            linha.append(cell)
        linhas_modelo.append(linha)

    wb_destino = openpyxl.load_workbook(excel_output2)
    for ws in wb_destino.worksheets:
        # Inserir 4 linhas no topo - 1 linha em branco para separar do cabeçalho
        ws.insert_rows(1, amount=4)

        # Concatenar e copiar para a coluna A das 3 primeiras linhas
        for i, linha in enumerate(linhas_modelo, start=1):
            valor1 = str(linha[0].value) if linha[0].value is not None else ""
            valor2 = str(linha[1].value) if len(linha) > 1 and linha[1].value is not None else ""
            cell_destino = ws.cell(row=i, column=1)
            cell_destino.value = f"{valor1} {valor2}".strip()  # Concatenar com espaço

    wb_destino.save(excel_output2)

    return

def concat(excel_output1, excel_output2, excel_final):
        # Abrir os dois arquivos
    wb1 = openpyxl.load_workbook(excel_output1)
    wb2 = openpyxl.load_workbook(excel_output2)


    for sheet_name in wb2.sheetnames:
        sheet2 = wb2[sheet_name]
        if sheet_name in wb1.sheetnames:
            new_name = f"{sheet_name}_copy"
        else:
            new_name = sheet_name
        new_sheet = wb1.create_sheet(title=new_name)

        # Copiar largura das colunas
        for col_letter, dim in sheet2.column_dimensions.items():
            new_sheet.column_dimensions[col_letter].width = dim.width

        # Copiar altura das linhas
        for row_idx, dim in sheet2.row_dimensions.items():
            new_sheet.row_dimensions[row_idx].height = dim.height

        # Copiar valores e estilos das células
        for row in sheet2.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)


    # Salvar o arquivo final
    wb1.save(excel_final)

    os.remove(excel_output1)
    os.remove(excel_output2)

    return
