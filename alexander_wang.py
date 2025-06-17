import pdfplumber
import pandas as pd
import sys
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
import fractions  # Para converter frações como '1/4' para float
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP
import re
import tempfile
import os



'''
A função pdf_to_excel lê o ficheiro pdf e converte-o para um ficheiro excel
As páginas de pdf que são convertidas para excel são aquelas que contém tabelas com medidas
'''
def pdf_to_excel(nome_pdf, excel_name):
    with pdfplumber.open(nome_pdf) as pdf:
        styles = []
        sample_sizes = []

        with pd.ExcelWriter(excel_name, engine='xlsxwriter') as writer:
            for i, page in enumerate(pdf.pages):
                tables = page.extract_tables({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines"
                })

                text = page.extract_text()
                if not tables or not text:
                    continue

                # Procurar a partir da tabela que contém "POM"
                start_copying = False
                collected_tables = []

                for table in tables:
                    df = pd.DataFrame(table).astype(str)

                    # Começa a copiar se encontrar "POM"
                    if not start_copying and df.apply(lambda row: row.str.contains("POM", case=False).any(), axis=1).any():
                        start_copying = True

                    if start_copying:
                        collected_tables.append(df)

                # Só escreve no Excel se encontrou o "POM"
                if collected_tables:
                    # Extrair style e sample size
                    style_match = re.search(r"Style\s+([^\s]+)", text)
                    styles.append(style_match.group(1) if style_match else "")

                    size_match = re.search(r"Sample Size\s+([^\s]+)", text)
                    sample_sizes.append(size_match.group(1) if size_match else "")

                    # Juntar os DataFrames
                    final_df = pd.DataFrame()
                    for df in collected_tables:
                        final_df = pd.concat([final_df, df, pd.DataFrame([[""] * len(df.columns)])], ignore_index=True)

                    final_df.to_excel(writer, sheet_name=f'Page_{i+1}', index=False, header=False)

        return styles, sample_sizes

'''
A função inches_to_cm converte os valores em polegadas para cm
'''
def inches_to_cm(value):
    """Converte polegadas para centímetros, mantendo texto/células vazias"""
    try:
        if isinstance(value, str):
            value = value.strip()  # Remover espaços extras

            # Se for fração (ex: '-1 1/2'), processar corretamente
            if ' ' in value and '/' in value:  # Caso tenha parte inteira e fração
                parts = value.split()
                inteiro = int(parts[0])  # Parte inteira
                fracao = float(fractions.Fraction(parts[1]))  # Parte fracionária

                # Se o inteiro for negativo, a fração também deve ser negativa
                if inteiro < 0:
                    fracao = -fracao
                
                value = inteiro + fracao  # Soma as partes

            elif '/' in value:  # Apenas fração (ex: '1/4')
                value = float(fractions.Fraction(value))
            else:
                value = float(value)  # Se for número inteiro ou decimal
            
            cm_value = Decimal(str(value * 2.54)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return float(cm_value)  # Converter para cm e arredondar
        else:
            return value  # Manter outros valores intactos
    except:
        return value  # Se der erro, manter como está


'''
A função convert_selected_columns lê o ficheiro excel que foi criado pela função pdf_to_excel e faz a conversão de polegadas para centímetros e 
nas tabelas que contém medidas para tamanhos diferentes, calcula a diferença entre tamanhos consecutivos
'''
def convert_selected_columns(excel_entrada,excel_saida):
    xls = pd.ExcelFile(excel_entrada)

    with pd.ExcelWriter(excel_saida, engine='openpyxl', mode='w') as writer:
        for sheet_name in xls.sheet_names:

            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

            # Identificar colunas que contêm pelo menos um número ou fração e não estão totalmente vazias
            colunas_para_converter = []
            for coluna in df.columns:
                if "POM" in coluna:  # Ignorar qualquer coluna que contenha "POM" no nome
                    continue  

                valores_nao_nulos = df[coluna].dropna().astype(str).str.strip()  # Remove NaN e espaços extras
                
                if valores_nao_nulos.empty:
                    continue  # Ignorar colunas completamente vazias
                
                #verificar se todos os valores não vazios na coluna são números ou frações, em caso afirmativo adicionar à lista colunas_para_converter
                #tem que se acrescentar .replace('-',"") para reconhecer numeros negativos como sendo digitos
                if valores_nao_nulos.apply(lambda x: x.replace(" ", "").replace("/", "").replace('-',"").isdigit()).all():
                    colunas_para_converter.append(coluna)

            # Criar DataFrame final, mantendo as colunas convertidas logo após as originais
            df_final = df.copy()
            
            for coluna in colunas_para_converter:
                nome_cm = f"{coluna} (cm)"
                df_final[nome_cm] = df[coluna].apply(inches_to_cm)

                # Reorganizar as colunas para manter a ordem original
                colunas = df_final.columns.tolist()
                idx = colunas.index(coluna)  # Pegar a posição da coluna original
                colunas.insert(idx + 1, colunas.pop(colunas.index(nome_cm)))  # Inserir a nova coluna logo após a original
                df_final = df_final[colunas]  # Atualizar a ordem das colunas

            # Para calcular a diferença entre os tamanhos:
            # Selecionar as colunas com tamnahos que estão em cm
            colunas = df_final.columns.tolist()
            SIZES = ['XXS (cm)', 'XS (cm)', '[S] (cm)', 'S (cm)', 'M (cm)', 'L (cm)', 'XL (cm)', 'XXL (cm)','[XXS] (cm)','[XS] (cm)',
                     '[M] (cm)','[L] (cm)','[XL] (cm)','[XXL] (cm)']
            # Selecionar apenas as colunas que estão na lista de tamanhos
            colunas_cm = [col for col in colunas if col in SIZES]
            colunas_diff = []
            if colunas_cm != []:
                # Loop para calcular diferenças entre colunas (cm) e inserir a nova coluna entre as duas colunas
                # que foram usadas para calcular a diferença
                for i in range(1, len(colunas_cm)):
                    # Selecionar a coluna atual e a anterior a esta para ser calculada a diferença entre as duas
                    coluna_atual = colunas_cm[i]
                    coluna_anterior = colunas_cm[i - 1]

                    # Nome da nova coluna de diferença
                    nova_coluna = f'Dif {coluna_atual[:-5]}-{coluna_anterior[:-5]}'
                    colunas_diff.append(nova_coluna)

                    # Calcular a diferença
                    df_final[nova_coluna] = df_final[coluna_atual] - df_final[coluna_anterior]
                    # Encontrar o índice da coluna anterior para inserir a coluna da diferença após esta coluna
                    indice_coluna = df_final.columns.get_loc(coluna_anterior)
                    colunas_atualizadas = df_final.columns.tolist()

                    # Remover a nova coluna do final e inserir após a coluna anterior
                    colunas_atualizadas.remove(nova_coluna)
                    colunas_atualizadas.insert(indice_coluna + 1, nova_coluna)

                    # Atualizar a ordem das colunas no DataFrame
                    df_final = df_final[colunas_atualizadas]

            # renomear as colunas com diferença entre tamanhos por "":
            novo_nome = {col: "" for col in colunas_diff if col in df_final.columns}
            df = df_final.rename(columns=novo_nome)
            df_final=df        

            # Salvar no Excel
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)
    xls.close()


'''
A função formatar_excel formata algumas definições do excel, como ajustar a largura e altura de alguma celulas e acrescentar limites nas celulas
'''
def formatar_excel(nome_excel):
    workbook = openpyxl.load_workbook(nome_excel)
    for sheet in workbook.worksheets:
        # Definir a largura da primeira coluna (A)
        sheet.column_dimensions['A'].width = 12
        # Definir a largura da primeira coluna (B)
        sheet.column_dimensions['B'].width = 45


        start_col_idx = openpyxl.utils.column_index_from_string('D')

        # Iterar por todas as colunas a partir de 'D' e ajustar a largura
        for col_idx in range(start_col_idx, sheet.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 7

        # Ajustar a altura das linhas e fazer o texto se ajustar ao tamanho da célula
        #coloquei até 100 para ter a certeza que aplica a altura das celulas às linhas todas
        for row_idx in range(1, 100):
            sheet.row_dimensions[row_idx].height = 20# Ajuste a altura conforme necessário

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                #if cell.column >= start_col_idx:  # Apenas células a partir da coluna 'D'
                #    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula

    # Salvar o arquivo Excel
    workbook.save(nome_excel)
    workbook.close()

    wb = load_workbook(nome_excel)
    ws = wb.active
    
    for sheet in wb.worksheets:
        # Iterar sobre as células da primeira linha (cabeçalho) e aplicar wrap_text e centralizar
        for cell in sheet[1]:  # sheet[1] acessa a primeira linha
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Ajustar a altura da primeira linha para garantir que o texto quebre
        sheet.row_dimensions[5].height = 25  # Ajuste conforme necessário

    # Salvar as alterações
    wb.save(nome_excel)
    wb.close()

#a partir da coluna D se tiver apenas zeros estes vao ser removidos, visto que não acrescentam informação
# ...existing code...
def remove_zeros(nome_excel):
    workbook = openpyxl.load_workbook(nome_excel)
    for sheet in workbook.worksheets:
        start_col_idx = openpyxl.utils.column_index_from_string('D')
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            cells_from_d = row[start_col_idx-1:]
            # Considere todos os tipos de "zero" ou vazio
            if all((cell.value is None) or(str(cell.value).strip() in ("", "0", "None"))for cell in cells_from_d):
                for cell in cells_from_d:
                    cell.value = ""
    workbook.save(nome_excel)
    workbook.close()


def add_info(excel_output2, lista_styles, lista_sample_sizes):
    wb_destino = openpyxl.load_workbook(excel_output2)
    i=0
    for ws in wb_destino.worksheets:
        # Inserir 4 linhas no topo - 1 linha em branco para separar do cabeçalho
        ws.insert_rows(1, amount=4)
        valor1 = str(lista_styles[i]) 
        valor2 = str(lista_sample_sizes[i])
        cell_style = ws.cell(row=2, column=1)
        cell_style.value = "STYLE:"
        cell_style1 = ws.cell(row=2, column=2)
        cell_style1.value = f"{valor1}".strip()  # Concatenar com espaço
        cell_style.font = Font(bold=True)  # Deixa em negrito
     
        cell_ssize = ws.cell(row=3, column=1)
        cell_ssize.value = "SAMPLE SIZE:"
        cell_ssize1 = ws.cell(row=3, column=2)
        cell_ssize1.value = f"{valor2}".strip()  # Concatenar com espaço
        cell_ssize.font = Font(bold=True)  # Deixa em negrito

        i=+1
    wb_destino.save(excel_output2)

    return

"RUN CODE"
if __name__ == "__main__":
    #carregar ficheiro pdf
    uploaded_pdf = st.file_uploader("Carrega o PDF")
    #definir nome do ficheiro pdf
    pdf_name = f'z_processed_files/{sys.argv[1]}.pdf'
    #definir o nome do ficheiro excel para o qual será transferida a informação do pdf
    excel_entrada = f'z_processed_files/{sys.argv[1]}.xlsx'
    #definir o nome do ficheiro excel que irá conter as alterações: conversão para cm e calculo da diferença entre tamanhos
    excel_saida = f'z_processed_files/{sys.argv[1]}_processed.xlsx'

    #conversão de pdf para excel
    styles, sample_sizes=pdf_to_excel(pdf_name,excel_entrada)


    #sys.exit()

    #processamento do excel criado
    convert_selected_columns(excel_entrada,excel_saida)



    #formatar ficheiro excel
    #formatar_excel(excel_saida)

    #Remover o primeiro ficheiro excel criado uma vez que já não será utilizado
    os.remove(excel_entrada)

    #sys.exit()

    # Chama a função para formatar o arquivo
    formatar_excel(excel_saida)

    remove_zeros(excel_saida)

    "acrescentar STYLE e SAMPLE SIZE"
    add_info(excel_saida, styles, sample_sizes)


    print('                     PROCESSO TERMINADO                ')
    print('                          :)                            ')
    print(f'    VER FICHEIRO -----> {sys.argv[1]}_processed.xlsx     ')



