import pandas as pd
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side
import fractions  # Para converter frações como '1/4' para float
from openpyxl import load_workbook
from fractions import Fraction
import re
from decimal import Decimal, ROUND_HALF_UP


'''
A função inches_to_cm converte os valores em polegadas para cm
'''
def inches_to_cm(value):
    """Converte polegadas para centímetros, mantendo texto/células vazias"""
    try:
        if isinstance(value, str):
            value = value.strip()  # Remover espaços extras

            # Se for fração (ex: '-1 1/2'), processar corretamente
            if ' ' in value and '/' in value:  # Caso tenha parte inteira e fração
                parts = value.split()
                inteiro = int(parts[0])  # Parte inteira
                fracao = float(fractions.Fraction(parts[1]))  # Parte fracionária

                # Se o inteiro for negativo, a fração também deve ser negativa
                if inteiro < 0:
                    fracao = -fracao
                
                value = inteiro + fracao  # Soma as partes

            elif '/' in value:  # Apenas fração (ex: '1/4')
                value = float(fractions.Fraction(value))
            else:
                value = float(value)  # Se for número inteiro ou decimal

            cm_value = Decimal(str(value * 2.54)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            return float(cm_value)
        else:
            return value  # Manter outros valores intactos
    except:
        return value  # Se der erro, manter como está
    
'''
A função decimal_para_fracao é usada apenas para colocar os valores em polegadas na formataçõ original, ao ler o ficheiro excel os valores
em frações são convertidos para decimal e apenas para manter a formatação original é feita a converção em fração usando esta função
'''
def decimal_para_fracao(valor):
    """Converte um número decimal para fração no formato inteiro e fração (ex: 20.75 -> 20 3/4)."""
  
    # Verifica se o número é negativo
    negativo = valor < 0
    valor = abs(valor)  # Trabalhamos com o valor positivo

    # Divide a parte inteira e a parte fracionária
    inteiro = int(valor)  # Parte inteira
    frac = valor - inteiro  # Parte decimal

    # Se não houver parte fracionária, retorna apenas o número inteiro
    if frac == 0:
        return f"-{inteiro}" if negativo else str(inteiro)

    # Converter a parte decimal para string para comparar os primeiros dígitos
    str_frac = str(round(frac, 5)).split('.')[1]  

    # Frações comuns e seus equivalentes decimais
    excecoes = {
        "25": (1, 4),
        "50": (1, 2),
        "75": (3, 4),
        "125": (1, 8),
        "375": (3, 8),
        "625": (5, 8),
        "875": (7, 8)
    }

    # Verifica os três primeiros dígitos primeiro (casos como 0.125, 0.375)
    if str_frac[:3] in excecoes:
        numerador, denominador = excecoes[str_frac[:3]]
    # Caso contrário, verifica os dois primeiros dígitos (casos como 0.25, 0.50, 0.75)
    elif str_frac[:2] in excecoes:
        numerador, denominador = excecoes[str_frac[:2]]
    else:
        n = len(str_frac)  # Número de casas decimais
        denominador = 10 ** n  # Denominador será 10^n
        numerador = int(frac * denominador)  # Multiplica a parte decimal pelo denominador
        
        # Simplificar a fração dividindo pelo 5, o máximo de vezes possível
        while numerador % 5 == 0 and denominador % 5 == 0:
            numerador //= 5
            denominador //= 5
    
    # Se o número for negativo, o numerador deve ser negativo
    if negativo:
        numerador = -numerador

    # Se a parte inteira for zero, apenas retorna a fração com o sinal correto
    if inteiro == 0:
        return f"{numerador}/{denominador}"
    else:
        # Se houver parte inteira, combina a parte inteira com a fração, mantendo o sinal correto
        return f"{'-' if negativo else ''}{inteiro} {abs(numerador)}/{denominador}"

'''
A função selecionar_tabelas seleciona apenas as sheets e tabelas pretendidas
'''
def selecionar_tabelas(excel_entrada,keywords,excel_saida):
    workbook=openpyxl.load_workbook(excel_entrada,data_only=True)
    sheet_names=workbook.sheetnames

    # Remover espaços em branco no início e no fim dos nomes das sheets
    sheets_to_keep = [sheet_name for sheet_name in sheet_names 
                        if any(keyword.lower() in sheet_name.strip().lower() for keyword in keywords)]

    # Remover apenas as sheets que não estão na lista de sheets a manter
    sheets_to_remove = [sheet_name for sheet_name in sheet_names if sheet_name not in sheets_to_keep]

        # Para cada aba (sheet) no arquivo
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (float, int)):  # Se o valor for numérico
                    # Converte o valor decimal para fração e coloca como texto na célula
                    cell.value = decimal_para_fracao(cell.value)

    for sheet_name in sheets_to_remove:
        std = workbook[sheet_name]
        workbook.remove(std)

    output_file = f'{sys.argv[1]}_aux.xlsx'
    workbook.save(output_file)
    workbook.close()
    
    xls = pd.ExcelFile(output_file)

    with pd.ExcelWriter(excel_saida,engine='xlsxwriter') as writer:
        for sheet_name in xls.sheet_names:
            df_new = pd.read_excel(xls, sheet_name=sheet_name, dtype=str, header=None)  # Ler como texto

            linha_pom = None  # Inicializa a variável
            # Encontrar a linha onde está "P.O.M"
            for i in range(len(df_new)):  
                if df_new.iloc[i].astype(str).str.contains("P.O.M", na=False).any():
                    linha_pom = i
                    break
            
            # Se "P.O.M" foi encontrado, remover linhas antes dele
            if linha_pom is not None:
                df_new = df_new.iloc[linha_pom:].reset_index(drop=True)

                # Definir a primeira linha como cabeçalho
                df_new.columns = df_new.iloc[0].astype(str)
                df_new = df_new[1:].reset_index(drop=True)

            # Remove espaços em branco extras no inicio e fim do nome das colunas
            df_new.columns = df_new.columns.str.strip()

            # Remover espaços extras dentro do nome das colunas (preservando o conteúdo)
            df_new.columns = [col.replace(" ", "") if "TOL" in col else col for col in df_new.columns]

            # Identificar e remover a coluna 'SKETCH' e 'nan'
            df_new = df_new.drop(columns=[col for col in df_new.columns if col.strip().upper() == 'SKETCH' or col.strip() == 'nan'])

            # Salvar o novo arquivo Excel
            df_new.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

    xls.close()

    os.remove(output_file)
    
'''
A função convert_selected_columns lê o ficheiro excel que foi criado pela função pdf_to_excel e faz a conversão de polegadas para centímetros e 
nas tabelas que contém medidas para tamanhos diferentes, calcula a diferença entre tamanhos consecutivos
'''
def convert_selected_columns(excel_saida):
    xls = pd.ExcelFile(excel_saida)

    with pd.ExcelWriter(excel_saida, engine='xlsxwriter') as writer:
        for sheet_name in xls.sheet_names:
            #ler ficheiro excel
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str)

            # Identificar colunas que contêm pelo menos um número ou fração e não estão totalmente vazias
            colunas_para_converter = []
            # Loop sobre as colunas do DataFrame
            for coluna in df.columns:
                valores_nao_nulos = df[coluna].dropna().astype(str).str.strip()  # Remover NaN e espaços extras
                
                # Verificar se ao menos uma célula na coluna é um número ou fração válida
                for valor in valores_nao_nulos:
                    valor = valor.strip()  # Remove espaços extras
                    if not valor:  # Se a célula estiver vazia, continua
                        continue

                    # Verificar se é um número (inteiro ou decimal) ou uma fração
                    if re.match(r'^\d+(\.\d+)?$', valor) or re.match(r'^\d+/\d+$', valor) or re.match(r'^\d+\s+\d+/\d+$', valor):
                        colunas_para_converter.append(coluna)

            
            # Criar DataFrame final, mantendo as colunas convertidas logo após as originais
            df_final = df.copy()
            
            for coluna in colunas_para_converter:
                nome_cm = f"{coluna} (cm)"
                df_final[nome_cm] = df[coluna].apply(inches_to_cm)

                # Reorganizar as colunas para manter a ordem original
                colunas = df_final.columns.tolist()
                idx = colunas.index(coluna)  # Pegar a posição da coluna original
                colunas.insert(idx + 1, colunas.pop(colunas.index(nome_cm)))  # Inserir a nova coluna logo após a original
                df_final = df_final[colunas]  # Atualizar a ordem das colunas
            
            # Para calcular a diferença entre os tamanhos:
            # Selecionar as colunas com tamnahos que estão em cm
            colunas = df_final.columns.tolist()
            SIZES = ['XXS (cm)', 'XS (cm)', '[S] (cm)', 'S (cm)', 'M (cm)', 'L (cm)', 'XL (cm)', 'XXL (cm)','[XXS] (cm)',
                     '[XS] (cm)','[M] (cm)','[L] (cm)','[XL] (cm)','[XXL] (cm)']
            # Selecionar apenas as colunas que estão na lista de tamanhos
            colunas_cm = [col for col in colunas if col in SIZES]
            if colunas_cm != []:
                # Loop para calcular diferenças entre colunas (cm) e inserir a nova coluna entre as duas colunas
                # que foram usadas para calcular a diferença
                for i in range(1, len(colunas_cm)):
                    # Selecionar a coluna atual e a anterior a esta para ser calculada a diferença entre as duas
                    coluna_atual = colunas_cm[i]
                    coluna_anterior = colunas_cm[i - 1]

                    # Nome da nova coluna de diferença
                    nova_coluna = f'Dif {coluna_atual[:-5]}-{coluna_anterior[:-5]}'

                    # Calcular a diferença
                    df_final[nova_coluna] = df_final[coluna_atual] - df_final[coluna_anterior]

                    # Encontrar o índice da coluna anterior para inserir a coluna da diferença após esta coluna
                    indice_coluna = df_final.columns.get_loc(coluna_anterior)
                    colunas_atualizadas = df_final.columns.tolist()

                    # Remover a nova coluna do final e inserir após a coluna anterior
                    colunas_atualizadas.remove(nova_coluna)
                    colunas_atualizadas.insert(indice_coluna + 1, nova_coluna)

                    # Atualizar a ordem das colunas no DataFrame
                    df_final = df_final[colunas_atualizadas]
            
            
            # Salvar no Excel
            df_final.to_excel(writer, sheet_name=sheet_name, index=False)

    xls.close()

'''
A função formatar_excel formata algumas definições do excel, como ajustar a largura e altura de alguma celulas e acrescentar limites nas celulas
'''
def formatar_excel(nome_excel):
    workbook = openpyxl.load_workbook(nome_excel)
    for sheet in workbook.worksheets:
        # Definir a largura da primeira coluna (A)
        sheet.column_dimensions['A'].width = 65

        start_col_idx = openpyxl.utils.column_index_from_string('B')

        # Iterar por todas as colunas a partir de 'D' e ajustar a largura
        for col_idx in range(start_col_idx, sheet.max_column + 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 10

        # Ajustar a altura das linhas e fazer o texto se ajustar ao tamanho da célula
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                sheet.row_dimensions[cell.row].height = 20  # Ajuste a altura conforme necessário

        for row in sheet.iter_rows():
            for cell in row:
                if cell.column >= start_col_idx:  # Apenas células a partir da coluna 'D'
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula

    # Salvar o arquivo Excel
    workbook.save(nome_excel)
    workbook.close()

    #moldar o texto da primeira linha do excel
    wb = load_workbook(nome_excel)
    ws = wb.active
    
    for sheet in wb.worksheets:
        # Iterar sobre as células da primeira linha (cabeçalho) e aplicar wrap_text e centralizar
        for cell in sheet[1]:  # sheet[1] acessa a primeira linha
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Ajustar a altura da primeira linha para garantir que o texto quebre
        sheet.row_dimensions[1].height = 25  # Ajuste conforme necessário

    # Salvar as alterações
    wb.save(nome_excel)
    wb.close()

