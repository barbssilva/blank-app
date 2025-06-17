import pdfplumber
import pandas as pd
import sys
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl import load_workbook
import numpy as np
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage

from PIL import Image
import io


def pdf_to_excel(nome_pdf,excel_name):
    with pdfplumber.open(nome_pdf) as pdf:
        tables = []

        #esta parte serve apenas para guardar o modelo referente ao ficheiro pdf
        first_page = pdf.pages[0]  # Pegar apenas a primeira página
        # Extrair texto da primeira página
        text = first_page.extract_text()
        if text:  # Verifica se o texto foi extraído corretamente
            text_lines = text.split("\n")  # Divide o texto por linha
            text_df = pd.DataFrame(text_lines)  # Cria o DataFrame corretamente

        for page in pdf.pages:
            table = page.extract_table()
            table = [[cell.replace("\n", " ") if isinstance(cell, str) else cell for cell in row] for row in table]
            if table:
                # Convert table to DataFrame
                df = pd.DataFrame(table)

                tables.append(df)

            if tables:
                final_df = pd.concat(tables, ignore_index=True)
                final_df.to_excel(excel_name, index=False, header=False)

    return text_df

def excel_processing(excel_entrada, excel_saida):
    xls = pd.ExcelFile(excel_entrada)
    with pd.ExcelWriter(excel_saida, engine='xlsxwriter') as writer:
        df = pd.read_excel(xls,header = None)

        # Remover duplicatas a partir da terceira linha
        df_final = df.drop_duplicates()

        #eliminar linhas que dizem Visualizzazione risultati
        indices = df_final[df_final.apply(lambda row: row.astype(str).str.contains("Visualizzazione risultati", case=False, na=False)).any(axis=1)].index
        df_final = df_final.drop(indices)

        header = df_final.iloc[0]  # Pega a primeira linha como nomes das colunas
        df_final = df_final[1:]  # Remove a primeira linha dos dados
        df_final.columns = header  # Define a nova linha como cabeçalho

        # Salvar no Excel atualizado
        df_final.to_excel(writer, index=False,header=True)
    xls.close()

def dif_calc(excel_saida):
    xls = pd.ExcelFile(excel_saida)
    with pd.ExcelWriter(excel_saida, engine='xlsxwriter') as writer:
        df_final = pd.read_excel(xls)

        colunas_apenas_numeros = {}

            # Percorrer todas as colunas
        for coluna in df_final.columns:
                    # Ignorar a primeira linha da coluna
            valores = df_final[coluna].iloc[1:]
                    
                    # Remover NaN
            valores = valores.dropna()
                    
            if valores.empty:
                colunas_apenas_numeros[coluna] = False
                continue

            # Verificar se todos os valores restantes são numéricos (considerando ',' como separador decimal)
            if valores.apply(lambda x: str(x).replace('.', '').replace(',', '', 1).isdigit()).all():
                colunas_apenas_numeros[coluna] = True
            else:
                colunas_apenas_numeros[coluna] = False

                # Obter a lista de colunas que contêm apenas números
        colunas_numericas = [col for col, is_numeric in colunas_apenas_numeros.items() if is_numeric]

        # Remover linhas que contêm 'Incremento'
        indices_incremento = df_final[df_final.apply(lambda row: row.astype(str).str.contains("Incremento", case=False, na=False)).any(axis=1)].index
        df_final = df_final.drop(indices_incremento)

        for col in colunas_numericas: 
            df_final[col] = df_final[col].astype(str).str.replace(",", ".")  # Troca , por .
    
            df_final.loc[:,col] = pd.to_numeric(df_final[col], errors="coerce")  # Converte para número  # Converte para número

        novas_colunas =[]
        for i in range(1, len(colunas_numericas)):
            coluna_atual = colunas_numericas[i]
            coluna_anterior = colunas_numericas[i-1]

            nova_coluna = f'Dif {coluna_atual}-{coluna_anterior}'
            novas_colunas.append(nova_coluna)

                # Calcular a diferença
                #df_final[nova_coluna] = df_final[coluna_atual] - df_final[coluna_anterior]
                # Criar uma nova coluna para a diferença, preservando a primeira linha
            #df_final[nova_coluna] = np.nan
            #df_final.loc[1:,nova_coluna] = (df_final.loc[1:,coluna_atual] - df_final.loc[1:,coluna_anterior]).astype(float)

            df_final[nova_coluna] = df_final[coluna_atual] - df_final[coluna_anterior]  
            indice_coluna = df_final.columns.get_loc(coluna_anterior)
            colunas_atualizadas = df_final.columns.tolist()

            # Remover a nova coluna do final e inserir após a coluna anterior
            colunas_atualizadas.remove(nova_coluna)
            colunas_atualizadas.insert(indice_coluna + 1, nova_coluna)

            # Atualizar a ordem das colunas no DataFrame
            df_final = df_final[colunas_atualizadas]
        for col in novas_colunas:
            # Alterar o nome da nova coluna para ""
            df_final.rename(columns={col: ""}, inplace=True)

        # Salvar no Excel atualizado
        df_final.to_excel(writer, index=False,header=True)
    xls.close()

def formatar_excel(nome_excel):
    workbook = openpyxl.load_workbook(nome_excel)
    ws = workbook.active  # Usa a primeira sheet
    for sheet in workbook.worksheets:
        # Definir a largura da primeira coluna (B)
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 35


        start_col_idx = openpyxl.utils.column_index_from_string('E')

        # Ajustar a altura das linhas e fazer o texto se ajustar ao tamanho da célula
        for row in sheet.iter_rows(min_col=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                sheet.row_dimensions[cell.row].height = 40  # Ajuste a altura conforme necessário

        for row in sheet.iter_rows():
            for cell in row:
                if cell.column >= start_col_idx:  # Apenas células a partir da coluna 'E'
                    col_letter = get_column_letter(cell.column)
                    sheet.column_dimensions[col_letter].width = 5
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Definir a espessura da borda
        border_style = Border(
            left=Side(style='thin', color='000000'),  # Bordas à esquerda
            right=Side(style='thin', color='000000'),  # Bordas à direita
            top=Side(style='thin', color='000000'),  # Bordas em cima
            bottom=Side(style='thin', color='000000')  # Bordas embaixo
        )

        # Iterar sobre todas as linhas e colunas da planilha
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border_style  # Definir a borda da célula

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=3):  # Colunas A e B
            cell_a, cell_b = row  # Pega as células da linha
            if cell_a.value and cell_b.value:  # Verifica se ambas têm valores
                if isinstance(cell_a.value, str) and isinstance(cell_b.value, str):  # Só para strings
                    if cell_a.value in cell_b.value:  # Se o texto da A estiver na B
                        cell_b.value = cell_b.value.replace(cell_a.value, cell_a.value + "\n")  # Adiciona \n sem remover nada

    # Salvar o arquivo Excel
    workbook.save(nome_excel)
    workbook.close()


def add_images(pdf_path,excel_path,inf_texto):
    # Carregar o arquivo Excel existente
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Inserir linhas vazias no topo (para espaço das imagens)
    ws.insert_rows(1, 8)

        # Adicionar texto no topo
    i=1
    for texto in inf_texto:
        cell=ws.cell(row=i, column=1, value=texto)  # Insere o texto na coluna A
        cell.font = Font(bold=True, size=14) 
        i+=1

    image_paths=[]
    with pdfplumber.open(pdf_path) as pdf:
        row, col = 3, 2  # Linha inicial e coluna inicial

        for i, page in enumerate(pdf.pages):
            images = page.images  # Obtém as imagens da página
            for j, img in enumerate(images):
                # Extrair a imagem
                img_data = img["stream"].get_data()
                image = Image.open(io.BytesIO(img_data))

                # Calcular nova dimensão mantendo a proporção
                new_width = int(image.width * 0.5)
                new_height = int(image.height * 0.5)
                image = image.resize((new_width, new_height))

                # Salvar como arquivo temporário
                img_path = f"temp_img_{i}_{j}.png"
                image.save(img_path)
                image_paths.append(img_path)    

                # Adicionar ao Excel
                excel_img = ExcelImage(img_path)
                ws.add_image(excel_img, f"{openpyxl.utils.get_column_letter(col)}{row}")

                # Atualizar posição para colocar lado a lado
                col += 2  # Avança duas colunas para espaçamento
                if col > 6:  # Se passar da coluna F, volta para A e desce uma linha
                    col = 1
                    row += 15

    # Salvar o arquivo Excel atualizado
    wb.save(excel_path)
    # Remover os ficheiros das imagens após inserir no Excel
    for img_path in image_paths:
        os.remove(img_path)
